#! /usr/bin/env python3
# multiplicationtable.py - takes a number N from the command line and creates an NxN multiplication table in an Excel spreadsheet.

import sys, openpyxl
from openpyxl.styles import Font

# create variable for Bold Font
boldObj = Font(bold=True)

# make N = to sys Argv 1
N = int(sys.argv[1])

# create workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Loop through rows and columns filling in numbers/ formulas
for i in range(1, N+1):
    sheet['A' + str(i)] = i
    sheet['A' + str(i)].font = boldObj
    sheet.cell(row=1, column=i).value = i
    sheet.cell(row=1, column=i).font = boldObj

for row in range(1, N+1):
    for column in range(1, N+1):
        answer = row*column
        sheet.cell(row=row, column=column).value = answer




# save workbook
wb.save('multiplicationTable.xlsx')
print('Saving Table to Excel File')
